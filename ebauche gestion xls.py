



import openpyxl, pyautogui
####ajout de détails sans importance pour essai de modif git####
liste_echantillons ='/home/harold/essaiGit/SAAF/Liste Echantillons_modifiée2.xlsx'

echantillon = pyautogui.prompt(text='Entrer le nom de l\'échnatillon à passer', title='Echantillon à passer' , default='echantillon')
#pyautogui.prompt(text='allo', title='hello' , default='defaut')#
#pyautogui.prompt(text='text', title='title', default='default')

print (echantillon)
wb = openpyxl.load_workbook(liste_echantillons)
sheet=wb.get_sheet_by_name('Feuil1')
for row in range (2, sheet.max_row+1):
    value = sheet['A' + str(row)].value
    #print row, value, type(echantillon), type(value)
    if value != echantillon:
        continue
    ligne = row
    print 'l\'echantillon', echantillon, ' a été trouvé à la ligne ', row


Profil = sheet['P' + str(ligne)].value
Protocole = sheet['Q' + str(ligne)].value

print pyautogui.prompt(text=('Profil et protocole de test', Profil, Protocole), title=('echantillon ',echantillon) , default=(Profil, Protocole))
debitAir = sheet['M' + str(ligne)].value
print 'debit d\'air ', debitAir, ' g/s'
pyautogui.alert(text=(debitAir, ' g/s'), title='Débit d\'air de refroidissement', button='OK')

base_fichier_sauvegarde = value = sheet['B' + str(ligne)].value
fichier_save = base_fichier_sauvegarde + '_C1-70'
print fichier_save, 
sheet.cell(row=int(ligne), column=18).value = fichier_save

wb.save(liste_echantillons)
